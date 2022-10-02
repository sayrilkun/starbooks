from importlib.resources import read_binary
import pandas as pd  # pip install pandas openpyxl
import plotly.express as px  # pip install plotly-express
import streamlit as st  # pip install streamlit
from openpyxl import load_workbook
import numpy as np
from sklearn.linear_model import LinearRegression
from PIL import Image
import webbrowser
# emojis: https://www.webfx.com/tools/emoji-cheat-sheet/
st.set_page_config(page_title="Median Lethal Dose Calculator", page_icon=":seedling:", layout="wide")
st.title(":seedling: Median Lethal Dose Calculator")
st.markdown("""---""")

# ---- READ EXCEL ----
@st.cache
def get_data(file,sheet_name):
    wb = load_workbook(filename=file, 
                   read_only=True,
                   data_only=True)
    ws = wb[sheet_name]
    dimension = ws.calculate_dimension()

    # Read the cell values into a list of lists
    data_rows = []
    for row in ws[dimension]:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)

    # # Transform into dataframe
    # import pandas as pd
    df = pd.DataFrame(data_rows)
    df['Average'] = df.mean(axis=1)
    df['%Germination'] = (df['Average']/25)*100

    return df


# ---- SIDEBAR ----

pnri = Image.open('logo-pnri.jpg')
st.sidebar.image(pnri)
# st.sidebar.header("Philippine Nuclear Reasearch Institute")
    # st.write("Philippine Nuclear Reasearch Institute")


mort_list = []
dose = []
st.sidebar.markdown("##")

st.sidebar.warning('This calculator requires a specific template of data from an excel file. Please download the sample file below for your guidance.')
# st.sidebar.download_button('sample.xlsx', 'sample.xlsx')
# if st.sidebar.button('Download Sample File'):
    # webbrowser.open('https://firebasestorage.googleapis.com/v0/b/pnri-demeter.appspot.com/o/sample.xlsx?alt=media&token=de387956-95b8-4a81-b5b2-fb25b37958eb')
with open("sample.xlsx", "rb") as file:
    btn = st.sidebar.download_button(
            label="Download Sample File",
            data=file,
            file_name="sample.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
uploaded_file = st.sidebar.file_uploader("Choose a file")
# if uploaded_file is not None:
    # df = get_data_from_excel(uploaded_file)
    # df = get_data(uploaded_file)
# sheet_name = st.sidebar.text_input('Sheet name')

# control = st.sidebar.text_input(
#     'Control Group Sheet Name', 
#     placeholder='Ex: Control Group')

# try:
if uploaded_file != None:
    # add_name = "Control Group,"+control
    # x = add_name.split(",")
    worbi = load_workbook(filename=uploaded_file)

    df = get_data(uploaded_file,"Control Group")
    st.subheader("🧫 Control Group")
    load_data("Control Group")

# except Exception as e:
#     st.sidebar.error("Invalid Input!")

# num_exp  = st.sidebar.text_input('Number of Experimental Group')

# # try:
# if num_exp != "":
    st.markdown("##")
    st.subheader("☢️ Experimental Group")
    for i in range(1,len(worbi.sheetnames)):
        # i = st.sidebar.text_input(
        #     f'{i+1}. Name of Experimental Group', 
        #     placeholder='Ex: 350 Gy')
        # if i != "":
        x = worbi.sheetnames[i].split(" ")
        print(x[0])
        df = get_data(uploaded_file,worbi.sheetnames[i])
        st.subheader(f'{x[0]} {x[1]}')
        load_data(x[0])

# except Exception as e:
#     st.sidebar.error("Invalid Input!")


if st.sidebar.button("Calculate LD50"):
    # st.subheader("Final Output")
    st.markdown("##")
    st.markdown("""---""")

    st.subheader("✔️ Corrected Mortality")
    st.info('If the mortality rate in control is greater than or equal to 10% proceed with correction of mortality rate by using the equation below.')
    moor_img = Image.open('morty.jpg')
    l,m,r = st.columns(3)
    with m:
        st.image(moor_img)
    st.markdown("##")

    mort_df = pd.DataFrame(mort_list)
    mort_df.columns =['Treatment', 'Mortality']
    cg_mort = mort_df['Mortality'][0]
    if cg_mort >=10:
        mort_df['Corrected % Mortality'] = normalize(cg_mort,mort_df['Mortality'])
        mort_df['Corrected % Mortality'][0] = "N/A"
    else:
        mort_df['Corrected % Mortality'] = "N/A"

    st.dataframe(mort_df)
    mort_df['Corrected % Mortality'][0] = 0

    st.markdown("##")
    st.subheader("🧬 Probit Analysis")

    probit_df = pd.DataFrame(mort_df)
    probit_df.columns = ['Dose (Gy)', 'Mortality', 'Corrected % Mortality']
    
    probit_df["Dose (Gy)"] = pd.to_numeric(probit_df["Dose (Gy)"].loc[1:], downcast="float")
    
    probit_df['Log Dose']= round(np.log10(probit_df['Dose (Gy)'].loc[1:]),2)
    # probit_df['Log Dose']= np.log10(probit_df['Dose (Gy)'].loc[1:])

# Referring to Probit Table
    if cg_mort >=10:
        b=probit_df['Corrected % Mortality'].values
    else:
        b=probit_df['Mortality'].values

    probit_val = []
    for i in b:
        probit_val.append(pTable_df['Value'][i])

    probit_df['Probits']= probit_val
    
    # probit_df['Log Dose']= probit_df['Log Dose'].apply(lambda x:round(x,2))
    probit_df['Corrected % Mortality'][0] = 'N/A'
    probit_df['Probits'][0] = 'N/A'
    probit_df['Dose (Gy)'][0] = 'Control Group'


    st.dataframe(probit_df)


    st.markdown("##")
    st.subheader("💉 Median Lethal Dose")


    left_column, right_column = st.columns(2)

    p_val = probit_df['Probits'].loc[1:].values
    l_val = probit_df['Log Dose'].loc[1:].values
    print(p_val)
    print(l_val)
    data = {'Log Dose': l_val, 'Probits': p_val}
    ld_df = pd.DataFrame(data)

    rows_count = len(df.index)

    # find LD50

    x1 = ld_df['Log Dose'][0]
    x2 = ld_df['Log Dose'].iloc[-1]

    y1 = ld_df['Probits'][0]
    y2 = ld_df['Probits'].iloc[-1]

    m  = (y1-y2)/(x1-x2)
    b  = (x1*y2 - x2*y1)/(x1-x2)

    # x= round((5-b)/m,2)
    no_exp= (5-b)/m
    x= round((5-b)/m,2)
    xe = np.ceil(10**x)

    with left_column:
        st.dataframe(ld_df)
    with right_column:
        eldi = f'''<H3>LD<sub>50</sub> ≈ {x}</H3>
        '''
        st.markdown(eldi,unsafe_allow_html=True)
        st.subheader(F"≈ {xe} Gy in {rows_count} Days")


    # with right_column:
    st.markdown("##")

    fig_ld_50 = px.line(
        ld_df,
        x= "Log Dose",
        y= "Probits",
        markers = True,
        color_discrete_sequence=["#000000"] * len(ld_df),
        template="plotly_white",

    )
    fig_ld_50.update_xaxes(showgrid=False, zeroline=False)
    fig_ld_50.update_yaxes(showgrid=False, zeroline=False)
    fig_ld_50.update_layout(plot_bgcolor="#FFFFFF")
    fig_ld_50.update_layout(
    title={
        'text': "Median Lethal Dose Graph",
        'y':1,
        'x':0.5,
        'xanchor': 'center',
        'yanchor': 'top'})

    # fig_ld_50.add_shape(type='line',
    #             x0=2.54,
    #             y0=5,
    #             x1=x,
    #             y1=5,
    #             line=dict(color='Red',),
    #             xref='x',
    #             yref='y'
    # )
    # fig_ld_50.add_shape(type='line',
    #             x0=x,
    #             y0=5,
    #             x1=x,
    #             y1=2.54,
    #             line=dict(color='Red',),
    #             xref='x',
    #             yref='y'
    # )
    st.plotly_chart(fig_ld_50, use_container_width=True)

    st.markdown("##")
    st.markdown("##")

    figgy = px.scatter(ld_df,         
            x= "Log Dose",
            y= "Probits", 
            trendline="ols",
            color_discrete_sequence=["#000000"] * len(ld_df),
            template="plotly_white",
            )
    figgy.update_xaxes(showgrid=False, zeroline=False)
    figgy.update_yaxes(showgrid=False, zeroline=False)
    figgy.update_layout(plot_bgcolor="#FFFFFF")
    figgy.update_layout(
    title={
        'text': "Line of Best Fit",
        'y':1,
        'x':0.5,
        'xanchor': 'center',
        'yanchor': 'top'})
    st.plotly_chart(figgy, use_container_width=True)

    # y = np.array(p_val)
    # # p_val = [ 3.45, 4.33, 5.31, 6.08, 6.48]
    # # l_val = [2.54,2.60,2.65,2.69,2.74]
    # x = np.array(l_val)
    # plt.plot(l_val, p_val)


    # m, b = np.polyfit(l_val, p_val, 1)

    # wey = x, m*l_val + b
    # plt.plot(wey)
    # plt.savefig('my_plot.png')
    # ploty = Image.open('my_plot.png')

    # l,m,r = st.columns(3)
    # with m:
    #     st.image(ploty)

    st.subheader("📈 Linear Regression Analysis")
    st.markdown("##")
    st.info('''Hypotheses:

        A. Null: 
        The variability in the mortality of seeds is not explained by the radiation dose.
    B. Alternative: 
        The variability in the mortality of seeds is explained by the radiation dose.
    ''')
    st.markdown("##")
    st.info('''Decision Rule:

        If the R-Squared value is positively significantly correlated, reject null hypothesis.
    ''')
    st.markdown("##")
    # frames = [probit_df]

    # result = pd.concat(frames,axis =1)
    # st.dataframe(result)
    lr()
# if st.button("Save as PDF"):
#     pdfkit.from_url('https://share.streamlit.io/sayrilkun/ld50-calculator/app.py','google.pdf')


# with open('sample.xlsx','rb') as f:
#     st.sidebar.download_button('Download Sample File', f) 
# ---- HIDE STREAMLIT STYLE ----
hide_st_style = """
            <style>
            MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            footer:after {
                content:'Made by Research Interns from National University Manila BSc Computer Engineering'; 
                visibility: visible;
                display: block;
                position: relative;
                #background-color: red;
                padding: 5px;
                top: 1px;
            }
            # header {visibility: hidden;}
            </style>
            """
st.markdown(hide_st_style, unsafe_allow_html=True)
